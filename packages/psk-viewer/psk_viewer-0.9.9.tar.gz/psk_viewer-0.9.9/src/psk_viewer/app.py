# -*- coding: utf-8 -*-
from __future__ import annotations

from numbers import Number
from pathlib import Path
from typing import Any, Callable, Final, Iterable, Sequence, cast

import numpy as np
import pandas as pd  # type: ignore
import pyqtgraph as pg  # type: ignore
from numpy.typing import NDArray
from pyqtgraph import GraphicsScene, PlotWidget
from pyqtgraph.GraphicsScene.mouseEvents import MouseClickEvent  # type: ignore
from pyqtgraph.exporters.ImageExporter import ImageExporter
from qtpy.QtCore import (
    QByteArray,
    QCoreApplication,
    QItemSelectionModel,
    QModelIndex,
    QPoint,
    QPointF,
    QRect,
    QRectF,
    Qt,
)
from qtpy.QtGui import QBrush, QCloseEvent, QGuiApplication, QPalette, QPen
from qtpy.QtWidgets import QAction, QMessageBox, QWidget

from .detection import correlation, peaks_positions
from .gui import GUI
from .plot_data_item import PlotDataItem
from .preferences import Preferences
from .utils import (
    all_cases,
    copy_to_clipboard,
    load_data_csv,
    load_data_fs,
    load_data_scandat,
    resource_path,
    superscript_number,
)

__all__ = ["App"]

_translate = QCoreApplication.translate

pg.ViewBox.suggestPadding = lambda *_: 0.0


def tick_strings(
    self: pg.AxisItem, values: Iterable[float], scale: float, spacing: float
) -> list[str]:
    """improve formatting of `AxisItem.tickStrings`"""

    if self.logMode:
        return cast(list[str], self.logTickStrings(values, scale, spacing))

    places: int = max(0, int(np.ceil(-np.log10(spacing * scale))))
    strings: list[str] = []
    v: float
    for v in values:
        vs: float = v * scale
        v_str: str
        if abs(vs) < 0.001 or abs(vs) >= 10000:
            v_str = f"{vs:g}".casefold()
            while "e-0" in v_str:
                v_str = v_str.replace("e-0", "e-")
            v_str = v_str.replace("+", "")
            if "e" in v_str:
                e_pos: int = v_str.find("e")
                man: str = v_str[:e_pos]
                exp: str = superscript_number(v_str[e_pos + 1 :])
                v_str = man + "×10" + exp
            v_str = v_str.replace("-", "−")
        else:
            v_str = f"{vs:0.{places}f}"
        strings.append(v_str)
    return strings


pg.AxisItem.tickStrings = tick_strings


class App(GUI):
    PSK_DATA_MODE: Final[int] = 1
    PSK_WITH_JUMP_DATA_MODE: Final[int] = 2
    FS_DATA_MODE: Final[int] = -1

    def __init__(
        self,
        file_path: Path | None = None,
        parent: QWidget | None = None,
        flags: Qt.WindowType = Qt.WindowType.Window,
    ) -> None:
        super().__init__(parent, flags)

        self._data_mode: int = 0

        self._is_dark: bool = (
            self.palette().color(QPalette.ColorRole.Window).lightness() < 128
        )

        self._ghost_line: pg.PlotDataItem = self.figure.plot(np.empty(0), name="")
        self._plot_line: pg.PlotDataItem = self.figure.plot(np.empty(0), name="")
        self._ghost_data: PlotDataItem = PlotDataItem()
        self._plot_data: PlotDataItem = PlotDataItem()

        self._ignore_scale_change: bool = False

        self.model_signal: NDArray[np.float_]
        try:
            self.model_signal = pd.read_csv(
                resource_path("averaged fs signal filtered.csv")
            ).values.ravel()
        except (OSError, BlockingIOError):
            self.model_signal = np.empty(0)
            self.box_find_lines.hide()
        self.box_find_lines.setDisabled(True)
        self.user_found_lines: pg.PlotDataItem = self._canvas.scatterPlot(
            np.empty(0), symbol="o", pxMode=True
        )
        self.automatically_found_lines: pg.PlotDataItem = self._canvas.scatterPlot(
            np.empty(0), symbol="o", pxMode=True
        )
        self.user_found_lines_data: NDArray[np.float_] = np.empty(0)
        self.automatically_found_lines_data: NDArray[np.float_] = np.empty(0)

        # cross-hair
        self._crosshair_v_line: pg.InfiniteLine = pg.InfiniteLine(
            angle=90, movable=False
        )
        self._crosshair_h_line: pg.InfiniteLine = pg.InfiniteLine(
            angle=0, movable=False
        )

        self._cursor_balloon: pg.TextItem = pg.TextItem(
            color="#ccc" if self._is_dark else "#333"
        )
        self.figure.addItem(self._cursor_balloon)

        self._mouse_moved_signal_proxy: pg.SignalProxy = pg.SignalProxy(
            cast(GraphicsScene, self.figure.scene()).sigMouseMoved, rateLimit=10, slot=self.on_mouse_moved
        )
        self._axis_range_changed_signal_proxy: pg.SignalProxy = pg.SignalProxy(
            self.figure.sigRangeChanged, rateLimit=20, slot=self.on_lim_changed
        )

        self.setup_ui()

        self.load_config()

        self.setup_ui_actions()

        if file_path is not None and file_path.exists():
            loaded: bool = self.load_data(file_path)
            self.toolbar.load_trace_action.setEnabled(loaded)
            if loaded:
                self.set_config_value("open", "location", file_path.parent)

    def setup_ui(self) -> None:
        ax: pg.AxisItem
        label: str
        if self._is_dark:
            self.figure.setBackground(pg.mkBrush(0, 0, 0))
            for label, ax_d in self._canvas.axes.items():
                ax = ax_d["item"]
                ax.setPen("d")
                ax.setTextPen("d")
        else:
            self.figure.setBackground(pg.mkBrush(255, 255, 255))
            for label, ax_d in self._canvas.axes.items():
                ax = ax_d["item"]
                ax.setPen("k")
                ax.setTextPen("k")

        self.figure.plotItem.addItem(self._crosshair_v_line, ignoreBounds=True)
        self.figure.plotItem.addItem(self._crosshair_h_line, ignoreBounds=True)
        self.hide_cursors()

        self.set_plot_line_appearance()
        self.set_marks_appearance()
        self.set_crosshair_lines_appearance()

        self.model_found_lines.fancy_table_numbers = self.settings.fancy_table_numbers
        self.model_found_lines.log10_gamma = self.settings.log10_gamma

        # customize menu
        titles_to_leave: list[str] = [
            self._canvas.ctrl.alphaGroup.parent().title(),
            self._canvas.ctrl.gridGroup.parent().title(),
        ]
        action: QAction
        for action in self._canvas.ctrlMenu.actions():
            if action.text() not in titles_to_leave:
                self._canvas.ctrlMenu.removeAction(action)
        self._canvas.vb.menu = self._canvas.ctrlMenu
        self._canvas.ctrlMenu = None
        self._canvas.vb.menu.addAction(self._view_all_action)
        self._canvas.ctrl.autoAlphaCheck.setChecked(False)
        self._canvas.ctrl.autoAlphaCheck.hide()
        self.figure.sceneObj.contextMenu = None

        self._install_translation()

    def closeEvent(self, event: QCloseEvent) -> None:
        """senseless joke in the loop"""
        close: QMessageBox = QMessageBox()
        close.setText(self.tr("Are you sure?"))
        close.setIcon(QMessageBox.Icon.Question)
        close.setWindowIcon(self.windowIcon())
        close.setWindowTitle(self.tr("Spectrometer Data Viewer"))
        close.setStandardButtons(
            QMessageBox.StandardButton.Yes
            | QMessageBox.StandardButton.No
            | QMessageBox.StandardButton.Cancel
        )
        close_code: int = (
            QMessageBox.StandardButton.No
            if self._plot_data.frequency_span > 0.0
            else QMessageBox.StandardButton.Yes
        )
        while close_code == QMessageBox.StandardButton.No:
            close_code = close.exec()

        if close_code == QMessageBox.StandardButton.Yes:
            self.settings.setValue("windowGeometry", self.saveGeometry())
            self.settings.setValue("windowState", self.saveState())
            self.settings.sync()
            event.accept()
        elif close_code == QMessageBox.StandardButton.Cancel:
            event.ignore()

    def load_config(self) -> None:
        self._loading = True
        # common settings
        if self.settings.contains("windowGeometry"):
            self.restoreGeometry(
                cast(QByteArray, self.settings.value("windowGeometry", QByteArray()))
            )
        else:
            app: QCoreApplication | None = QCoreApplication.instance()
            if isinstance(app, QGuiApplication):
                window_frame: QRect = self.frameGeometry()
                desktop_center: QPoint = (
                    app.primaryScreen().availableGeometry().center()
                )
                window_frame.moveCenter(desktop_center)
                self.move(window_frame.topLeft())
        self.restoreState(
            cast(QByteArray, self.settings.value("windowState", QByteArray()))
        )

        self.check_frequency_persists.setChecked(
            self.get_config_value("frequency", "persists", False, bool)
        )
        self.check_voltage_persists.setChecked(
            self.get_config_value("voltage", "persists", False, bool)
        )

        self.spin_threshold.setValue(
            self.get_config_value("lineSearch", "threshold", 12.0, float)
        )

        if (
            self.get_config_value("display", "unit", PlotDataItem.VOLTAGE_DATA, str)
            == PlotDataItem.GAMMA_DATA
        ):
            self._plot_data.data_type = PlotDataItem.GAMMA_DATA
        else:
            self._plot_data.data_type = PlotDataItem.VOLTAGE_DATA
        self.switch_data_action.setChecked(
            self._plot_data.data_type == PlotDataItem.GAMMA_DATA
        )
        self.display_gamma_or_voltage()

        self._loading = False
        return

    def setup_ui_actions(self) -> None:
        self.toolbar.open_action.triggered.connect(self.on_open_action_clicked)
        self.toolbar.clear_action.triggered.connect(self.clear)
        self.toolbar.open_ghost_action.triggered.connect(
            lambda: cast(None, self.load_ghost_data())
        )
        self.toolbar.clear_ghost_action.triggered.connect(self.clear_ghost)
        self.toolbar.differentiate_action.toggled.connect(
            self.calculate_second_derivative
        )
        self.toolbar.save_data_action.triggered.connect(self.save_data)
        self.toolbar.copy_figure_action.triggered.connect(self.copy_figure)
        self.toolbar.save_figure_action.triggered.connect(self.save_figure)
        self.toolbar.load_trace_action.triggered.connect(self.load_found_lines)
        self.toolbar.copy_trace_action.triggered.connect(self.copy_found_lines)
        self.toolbar.save_trace_action.triggered.connect(self.save_found_lines)
        self.toolbar.clear_trace_action.triggered.connect(self.clear_found_lines)
        self.toolbar.configure_action.triggered.connect(self.edit_parameters)

        self.spin_frequency_min.valueChanged.connect(self.spin_frequency_min_changed)
        self.spin_frequency_max.valueChanged.connect(self.spin_frequency_max_changed)
        self.spin_frequency_center.valueChanged.connect(
            self.spin_frequency_center_changed
        )
        self.spin_frequency_span.valueChanged.connect(self.spin_frequency_span_changed)
        self.button_zoom_x_out_coarse.clicked.connect(
            lambda: self.button_zoom_x_clicked(1.0 / 0.5)
        )
        self.button_zoom_x_out_fine.clicked.connect(
            lambda: self.button_zoom_x_clicked(1.0 / 0.9)
        )
        self.button_zoom_x_in_fine.clicked.connect(
            lambda: self.button_zoom_x_clicked(0.9)
        )
        self.button_zoom_x_in_coarse.clicked.connect(
            lambda: self.button_zoom_x_clicked(0.5)
        )
        self.button_move_x_left_coarse.clicked.connect(
            lambda: self.button_move_x_clicked(-500.0)
        )
        self.button_move_x_left_fine.clicked.connect(
            lambda: self.button_move_x_clicked(-50.0)
        )
        self.button_move_x_right_fine.clicked.connect(
            lambda: self.button_move_x_clicked(50.0)
        )
        self.button_move_x_right_coarse.clicked.connect(
            lambda: self.button_move_x_clicked(500.0)
        )
        self.check_frequency_persists.toggled.connect(
            self.check_frequency_persists_toggled
        )

        self.switch_data_action.toggled.connect(self.on_switch_data_action_toggled)
        self.spin_voltage_min.valueChanged.connect(self.spin_voltage_min_changed)
        self.spin_voltage_max.valueChanged.connect(self.spin_voltage_max_changed)
        self.button_zoom_y_out_coarse.clicked.connect(
            lambda: self.button_zoom_y_clicked(1.0 / 0.5)
        )
        self.button_zoom_y_out_fine.clicked.connect(
            lambda: self.button_zoom_y_clicked(1.0 / 0.9)
        )
        self.button_zoom_y_in_fine.clicked.connect(
            lambda: self.button_zoom_y_clicked(0.9)
        )
        self.button_zoom_y_in_coarse.clicked.connect(
            lambda: self.button_zoom_y_clicked(0.5)
        )
        self.check_voltage_persists.toggled.connect(
            self.on_check_voltage_persists_toggled
        )

        self.spin_threshold.valueChanged.connect(
            lambda new_value: self.set_config_value(
                "lineSearch", "threshold", new_value
            )
        )
        self.button_find_lines.clicked.connect(self.on_button_find_lines_clicked)
        self.button_clear_lines.clicked.connect(self.clear_automatically_found_lines)
        self.button_prev_line.clicked.connect(self.prev_found_line)
        self.button_next_line.clicked.connect(self.next_found_line)

        self.table_found_lines.doubleClicked.connect(self.on_table_cell_double_clicked)

        line: pg.PlotDataItem
        for line in (self.automatically_found_lines, self.user_found_lines):
            line.sigPointsClicked.connect(self.on_points_clicked)

        self._view_all_action.triggered.connect(
            lambda: cast(None, self._canvas.vb.autoRange(padding=0.0))
        )

        self.figure.sceneObj.sigMouseClicked.connect(self.on_plot_clicked)

    def on_xlim_changed(self, xlim: Iterable[float]) -> None:
        min_freq, max_freq = min(xlim), max(xlim)
        self._loading = True
        self.spin_frequency_min.setValue(min_freq)
        self.spin_frequency_max.setValue(max_freq)
        self.spin_frequency_span.setValue(max_freq - min_freq)
        self.spin_frequency_center.setValue(0.5 * (max_freq + min_freq))
        self.spin_frequency_min.setMaximum(max_freq)
        self.spin_frequency_max.setMinimum(min_freq)
        self._loading = False
        self.set_frequency_range(
            lower_value=self.spin_frequency_min.value(),
            upper_value=self.spin_frequency_max.value(),
        )

    def on_ylim_changed(self, ylim: Iterable[float | np.float_]) -> None:
        min_voltage, max_voltage = min(ylim), max(ylim)
        self._loading = True
        self.spin_voltage_min.setValue(min_voltage)
        self.spin_voltage_max.setValue(max_voltage)
        self.spin_voltage_min.setMaximum(max_voltage)
        self.spin_voltage_max.setMinimum(min_voltage)
        self._loading = False
        self.set_voltage_range(lower_value=min_voltage, upper_value=max_voltage)

    def on_points_clicked(
        self, item: pg.PlotDataItem, points: Iterable[pg.SpotItem], ev: MouseClickEvent
    ) -> None:
        if item.xData is None or item.yData is None:
            return
        if not self.trace_mode:
            return
        if ev.button() != Qt.MouseButton.LeftButton:
            return

        point: pg.SpotItem
        if ev.modifiers() == Qt.KeyboardModifier.ShiftModifier:
            items: NDArray[np.float_] = item.scatter.data["item"]
            index: NDArray[np.bool_] = np.full(items.shape, True, np.bool_)
            for point in points:
                index &= items != point
                self.automatically_found_lines_data = (
                    self.automatically_found_lines_data[
                        self.automatically_found_lines_data != point.pos().x()
                    ]
                )
                self.user_found_lines_data = self.user_found_lines_data[
                    self.user_found_lines_data != point.pos().x()
                ]

            item.setData(item.xData[index], item.yData[index])

            # update the table
            self.model_found_lines.set_lines(
                self._plot_data,
                (self.automatically_found_lines_data, self.user_found_lines_data),
            )

            self.toolbar.copy_trace_action.setEnabled(
                not self.model_found_lines.is_empty
            )
            self.toolbar.save_trace_action.setEnabled(
                not self.model_found_lines.is_empty
            )
            self.toolbar.clear_trace_action.setEnabled(
                not self.model_found_lines.is_empty
            )

        elif ev.modifiers() == Qt.KeyboardModifier.NoModifier:
            found_lines_frequencies: NDArray[np.float_] = (
                self.model_found_lines.all_data[:, 0]
            )
            selected_points: list[int] = [
                cast(int, np.argmin(np.abs(point.pos().x() - found_lines_frequencies)))
                for point in points
            ]
            self.on_points_selected(selected_points)

    def on_button_find_lines_clicked(self) -> None:
        self.status_bar.showMessage(
            f"Found {self.find_lines(self.spin_threshold.value())} lines"
        )

    def on_mouse_moved(self, event: tuple[QPointF]) -> None:
        if self._plot_line.xData is None and self._plot_line.yData is None:
            return
        pos: QPointF = event[0]
        if self.figure.sceneBoundingRect().contains(pos):
            point: QPointF = self._canvas.vb.mapSceneToView(pos)
            if self.figure.visibleRange().contains(point):
                self.status_bar.clearMessage()
                self._crosshair_v_line.setPos(point.x())
                self._crosshair_h_line.setPos(point.y())
                self._crosshair_h_line.setVisible(self.settings.show_crosshair)
                self._crosshair_v_line.setVisible(self.settings.show_crosshair)
                self._cursor_x.setVisible(True)
                self._cursor_y.setVisible(True)
                self._cursor_x.setValue(point.x())
                self._cursor_y.setValue(point.y())

                if self.settings.show_coordinates_at_crosshair:
                    self._cursor_balloon.setPos(point)
                    self._cursor_balloon.setHtml(
                        self._cursor_x.text() + "<br>" + self._cursor_y.text()
                    )
                    balloon_border: QRectF = self._cursor_balloon.boundingRect()
                    sx: float
                    sy: float
                    sx, sy = self._canvas.vb.viewPixelSize()
                    balloon_width: float = balloon_border.width() * sx
                    balloon_height: float = balloon_border.height() * sy
                    anchor_x: float = (
                        0.0
                        if point.x() - self.figure.visibleRange().left() < balloon_width
                        else 1.0
                    )
                    anchor_y: float = (
                        0.0
                        if self.figure.visibleRange().bottom() - point.y()
                        < balloon_height
                        else 1.0
                    )
                    self._cursor_balloon.setAnchor((anchor_x, anchor_y))
                self._cursor_balloon.setVisible(
                    self.settings.show_coordinates_at_crosshair
                )
            else:
                self.hide_cursors()
        else:
            self.hide_cursors()

    def on_plot_clicked(self, event: MouseClickEvent) -> None:
        pos: QPointF = event.scenePos()
        if not self.trace_mode:
            return
        if (
            event.modifiers() != Qt.KeyboardModifier.NoModifier
            or event.button() != Qt.MouseButton.LeftButton
        ):
            return
        if not self.figure.sceneBoundingRect().contains(pos):
            return
        x_span: float = cast(float, np.ptp(self._canvas.axes["bottom"]["item"].range))
        y_span: float = cast(float, np.ptp(self._canvas.axes["left"]["item"].range))
        point: QPointF = self._canvas.vb.mapSceneToView(pos)
        if self._plot_line.xData is None or not self._plot_line.xData.size:
            return
        distance: NDArray[np.float_] = np.min(
            np.hypot(
                (self._plot_line.xData - point.x()) / x_span,
                (self._plot_line.yData - point.y()) / y_span,
            )
        )
        if distance > 0.01:
            return
        closest_point_index: int = cast(
            int,
            np.argmin(
                np.hypot(
                    (self._plot_line.xData - point.x()) / x_span,
                    (self._plot_line.yData - point.y()) / y_span,
                )
            ),
        )

        # avoid the same point to be marked several times
        if (
            self.user_found_lines.xData is not None
            and self.user_found_lines.yData.size
            and np.any(
                (
                    self.user_found_lines.xData
                    == self._plot_line.xData[closest_point_index]
                )
                & (
                    self.user_found_lines.yData
                    == self._plot_line.yData[closest_point_index]
                )
            )
        ):
            return
        if (
            self.automatically_found_lines.xData is not None
            and self.automatically_found_lines.yData.size
            and np.any(
                (
                    self.automatically_found_lines.xData
                    == self._plot_line.xData[closest_point_index]
                )
                & (
                    self.automatically_found_lines.yData
                    == self._plot_line.yData[closest_point_index]
                )
            )
        ):
            return

        self.user_found_lines_data = np.append(
            self.user_found_lines_data, self._plot_line.xData[closest_point_index]
        )

        self.user_found_lines.setData(
            self.user_found_lines_data,
            self._plot_line.yData[
                self.model_found_lines.frequency_indices(
                    self._plot_data, self.user_found_lines_data
                )
            ],
        )

        self.model_found_lines.add_line(
            self._plot_data, self._plot_line.xData[closest_point_index]
        )
        if self.settings.copy_frequency:
            copy_to_clipboard(str(1e-6 * self._plot_line.xData[closest_point_index]))
        self.toolbar.copy_trace_action.setEnabled(True)
        self.toolbar.save_trace_action.setEnabled(True)
        self.toolbar.clear_trace_action.setEnabled(True)

    def on_lim_changed(self, arg: tuple[PlotWidget, list[list[float]]]) -> None:
        if self._ignore_scale_change:
            return
        rect: list[list[float]] = arg[1]
        xlim: list[float]
        ylim: list[float]
        xlim, ylim = rect
        self._ignore_scale_change = True
        self.on_xlim_changed(xlim)
        self.on_ylim_changed(ylim)
        self._ignore_scale_change = False

    def on_points_selected(self, rows: list[int]) -> None:
        self.table_found_lines.clearSelection()
        sm: QItemSelectionModel = self.table_found_lines.selectionModel()
        row: int
        for row in rows:
            index: QModelIndex = self.model_found_lines.index(row, 0)
            sm.select(
                index,
                QItemSelectionModel.SelectionFlag.Select
                | QItemSelectionModel.SelectionFlag.Rows,
            )
            self.table_found_lines.scrollTo(index)

    def spin_frequency_min_changed(self, new_value: float) -> None:
        if self._loading:
            return
        self._loading = True
        self.spin_frequency_max.setMinimum(new_value)
        self.spin_frequency_center.setValue(
            0.5 * (new_value + self.spin_frequency_max.value())
        )
        self.spin_frequency_span.setValue(self.spin_frequency_max.value() - new_value)
        self.set_frequency_range(
            lower_value=new_value, upper_value=self.spin_frequency_max.value()
        )
        self._loading = False

    def spin_frequency_max_changed(self, new_value: float) -> None:
        if self._loading:
            return
        self._loading = True
        self.spin_frequency_min.setMaximum(new_value)
        self.spin_frequency_center.setValue(
            0.5 * (self.spin_frequency_min.value() + new_value)
        )
        self.spin_frequency_span.setValue(new_value - self.spin_frequency_min.value())
        self.set_frequency_range(
            lower_value=self.spin_frequency_min.value(), upper_value=new_value
        )
        self._loading = False

    def spin_frequency_center_changed(self, new_value: float) -> None:
        if self._loading:
            return
        freq_span = self.spin_frequency_span.value()
        min_freq = new_value - 0.5 * freq_span
        max_freq = new_value + 0.5 * freq_span
        self._loading = True
        self.spin_frequency_min.setMaximum(max_freq)
        self.spin_frequency_max.setMinimum(min_freq)
        self.spin_frequency_min.setValue(min_freq)
        self.spin_frequency_max.setValue(max_freq)
        self.set_frequency_range(upper_value=max_freq, lower_value=min_freq)
        self._loading = False

    def spin_frequency_span_changed(self, new_value: float) -> None:
        if self._loading:
            return
        freq_center = self.spin_frequency_center.value()
        min_freq = freq_center - 0.5 * new_value
        max_freq = freq_center + 0.5 * new_value
        self._loading = True
        self.spin_frequency_min.setMaximum(max_freq)
        self.spin_frequency_max.setMinimum(min_freq)
        self.spin_frequency_min.setValue(min_freq)
        self.spin_frequency_max.setValue(max_freq)
        self.set_frequency_range(upper_value=max_freq, lower_value=min_freq)
        self._loading = False

    def button_zoom_x_clicked(self, factor: float) -> None:
        if self._loading:
            return
        freq_span = self.spin_frequency_span.value() * factor
        freq_center = self.spin_frequency_center.value()
        min_freq = freq_center - 0.5 * freq_span
        max_freq = freq_center + 0.5 * freq_span
        self._loading = True
        self.spin_frequency_min.setMaximum(max_freq)
        self.spin_frequency_max.setMinimum(min_freq)
        self.spin_frequency_min.setValue(min_freq)
        self.spin_frequency_max.setValue(max_freq)
        self.spin_frequency_span.setValue(freq_span)
        self.set_frequency_range(upper_value=max_freq, lower_value=min_freq)
        self._loading = False

    def button_move_x_clicked(self, shift: float) -> None:
        if self._loading:
            return
        freq_span = self.spin_frequency_span.value()
        freq_center = self.spin_frequency_center.value() + shift
        min_freq = freq_center - 0.5 * freq_span
        max_freq = freq_center + 0.5 * freq_span
        self._loading = True
        self.spin_frequency_min.setMaximum(max_freq)
        self.spin_frequency_max.setMinimum(min_freq)
        self.spin_frequency_min.setValue(min_freq)
        self.spin_frequency_max.setValue(max_freq)
        self.spin_frequency_center.setValue(freq_center)
        self.set_frequency_range(upper_value=max_freq, lower_value=min_freq)
        self._loading = False

    def check_frequency_persists_toggled(self, new_value: bool) -> None:
        if self._loading:
            return
        self.set_config_value("frequency", "persists", new_value)

    def spin_voltage_min_changed(self, new_value: float) -> None:
        if self._loading:
            return
        self._loading = True
        self.spin_voltage_max.setMinimum(new_value)
        self.set_voltage_range(
            lower_value=new_value, upper_value=self.spin_voltage_max.value()
        )
        self._loading = False

    def spin_voltage_max_changed(self, new_value: float) -> None:
        if self._loading:
            return
        self._loading = True
        self.spin_voltage_min.setMaximum(new_value)
        self.set_voltage_range(
            lower_value=self.spin_voltage_min.value(), upper_value=new_value
        )
        self._loading = False

    def button_zoom_y_clicked(self, factor: float) -> None:
        if self._loading:
            return
        min_voltage = self.spin_voltage_min.value()
        max_voltage = self.spin_voltage_max.value()
        voltage_span = abs(max_voltage - min_voltage) * factor
        voltage_center = (max_voltage + min_voltage) * 0.5
        min_voltage = voltage_center - 0.5 * voltage_span
        max_voltage = voltage_center + 0.5 * voltage_span
        self._loading = True
        self.spin_voltage_min.setMaximum(max_voltage)
        self.spin_voltage_max.setMinimum(min_voltage)
        self.spin_voltage_min.setValue(min_voltage)
        self.spin_voltage_max.setValue(max_voltage)
        self.set_voltage_range(upper_value=max_voltage, lower_value=min_voltage)
        self._loading = False

    def on_check_voltage_persists_toggled(self, new_value: bool) -> None:
        if self._loading:
            return
        self.set_config_value("voltage", "persists", new_value)

    def edit_parameters(self) -> None:
        preferences_dialog: Preferences = Preferences(self.settings, self)
        if preferences_dialog.exec() == Preferences.DialogCode.Rejected:
            return
        self._install_translation()
        self.set_plot_line_appearance()
        self.set_marks_appearance()
        self.set_crosshair_lines_appearance()
        self.model_found_lines.fancy_table_numbers = self.settings.fancy_table_numbers
        self.model_found_lines.log10_gamma = self.settings.log10_gamma
        if (
            self._data_mode == self.PSK_DATA_MODE
            and self._plot_data.frequency_span > 0.0
        ):
            jump: float = (
                round(self.settings.jump / self._plot_data.frequency_step)
                * self._plot_data.frequency_step
            )
            self.toolbar.differentiate_action.setEnabled(
                0.0 < jump < 0.25 * self._plot_data.frequency_span
            )
            if not (0.0 < jump < 0.25 * self._plot_data.frequency_span):
                self.toolbar.differentiate_action.blockSignals(True)
                self.toolbar.differentiate_action.setChecked(False)
                self.toolbar.differentiate_action.blockSignals(False)
        self.display_gamma_or_voltage()

    def hide_cursors(self) -> None:
        self._crosshair_h_line.setVisible(False)
        self._crosshair_v_line.setVisible(False)
        self._cursor_x.setVisible(False)
        self._cursor_y.setVisible(False)
        self._cursor_balloon.setVisible(False)

    def on_open_action_clicked(self) -> None:
        loaded: bool = self.load_data()
        self.toolbar.load_trace_action.setEnabled(
            loaded or self.toolbar.load_trace_action.isEnabled()
        )

    @property
    def line(self) -> PlotDataItem:
        return self._plot_line

    @property
    def label(self) -> str | None:
        return self._plot_line.name()

    def set_frequency_range(
        self, lower_value: float | np.float_, upper_value: float | np.float_
    ) -> None:
        self.figure.plotItem.setXRange(lower_value, upper_value, padding=0.0)

    def set_voltage_range(
        self, lower_value: float | np.float_, upper_value: float | np.float_
    ) -> None:
        self.figure.plotItem.setYRange(lower_value, upper_value, padding=0.0)

    def set_plot_line_appearance(self) -> None:
        self._plot_line.setPen(
            pg.mkPen(self.settings.line_color, width=0.5 * self.settings.line_thickness)
        )
        self._ghost_line.setPen(
            pg.mkPen(
                self.settings.ghost_line_color, width=0.5 * self.settings.line_thickness
            )
        )
        self._canvas.replot()

    def set_marks_appearance(self) -> None:
        pen: QPen = pg.mkPen(
            self.settings.mark_pen, width=0.5 * self.settings.mark_pen_thickness
        )
        brush: QBrush = pg.mkBrush(self.settings.mark_brush)
        self.automatically_found_lines.setSymbolPen(pen)
        self.automatically_found_lines.setSymbolBrush(brush)
        self.automatically_found_lines.setSymbolSize(self.settings.mark_size)
        self.user_found_lines.setSymbolPen(pen)
        self.user_found_lines.setSymbolBrush(brush)
        self.user_found_lines.setSymbolSize(self.settings.mark_size)
        self._canvas.replot()

    def set_crosshair_lines_appearance(self) -> None:
        pen: QPen = pg.mkPen(
            self.settings.crosshair_lines_color,
            width=0.5 * self.settings.crosshair_lines_thickness,
        )
        self._crosshair_v_line.setPen(pen)
        self._crosshair_h_line.setPen(pen)
        self._canvas.replot()

    def find_lines(self, threshold: float) -> int:
        if self._data_mode == 0 or self.model_signal.size < 2:
            return 0

        from scipy import interpolate  # type: ignore

        x: Final[NDArray[np.float_]] = self._plot_line.xData
        y: Final[NDArray[np.float_]] = self._plot_line.yData
        if x.size < 2 or y.size < 2:
            return 0

        found_lines: NDArray[np.float_]
        if self._data_mode == self.FS_DATA_MODE:
            # re-scale the signal to the actual frequency mesh
            x_model: NDArray[np.float_] = (
                np.arange(self.model_signal.size, dtype=x.dtype) * 0.1
            )
            interpol = interpolate.interp1d(x_model, self.model_signal, kind=2)
            x_model_new: NDArray[np.float_] = np.arange(
                x_model[0], x_model[-1], x[1] - x[0]
            )
            y_model_new: NDArray[np.float_] = interpol(x_model_new)
            found_lines = peaks_positions(
                x, correlation(y_model_new, x, y), threshold=1.0 / threshold
            )
        elif self._data_mode in (self.PSK_DATA_MODE, self.PSK_WITH_JUMP_DATA_MODE):
            found_lines = peaks_positions(x, y, threshold=1.0 / threshold)
        else:
            return 0

        self._ignore_scale_change = True
        if found_lines.size:
            self.automatically_found_lines_data = x[found_lines]
            self.automatically_found_lines.setData(x[found_lines], y[found_lines])
        else:
            self.automatically_found_lines.setData(np.empty(0), np.empty(0))
            self.automatically_found_lines_data = np.empty(0)

        # update the table
        self.model_found_lines.set_lines(
            self._plot_data,
            (self.automatically_found_lines_data, self.user_found_lines_data),
        )

        self.toolbar.copy_trace_action.setEnabled(not self.model_found_lines.is_empty)
        self.toolbar.save_trace_action.setEnabled(not self.model_found_lines.is_empty)
        self.toolbar.clear_trace_action.setEnabled(not self.model_found_lines.is_empty)

        self.button_clear_lines.setEnabled(bool(found_lines.size))
        self.button_next_line.setEnabled(bool(found_lines.size))
        self.button_prev_line.setEnabled(bool(found_lines.size))

        self._ignore_scale_change = False

        return found_lines.size

    def prev_found_line(self) -> None:
        if self.model_signal.size < 2:
            return

        init_frequency: float = self.spin_frequency_center.value()

        line_data: NDArray[np.float_] = self.automatically_found_lines.xData
        if line_data is None or not line_data.size:
            return
        i: int = cast(int, np.searchsorted(line_data, init_frequency, side="right") - 2)
        if 0 <= i < line_data.size and line_data[i] != init_frequency:
            self.spin_frequency_center.setValue(line_data[i])
            self.ensure_y_fits()

    def next_found_line(self) -> None:
        if self.model_signal.size < 2:
            return

        init_frequency: float = self.spin_frequency_center.value()

        line_data: NDArray[np.float_] = self.automatically_found_lines.xData
        if line_data is None or not line_data.size:
            return
        i: int = cast(int, np.searchsorted(line_data, init_frequency, side="left") + 1)
        if i < line_data.size and line_data[i] != init_frequency:
            self.spin_frequency_center.setValue(line_data[i])
            self.ensure_y_fits()

    def on_table_cell_double_clicked(self, index: QModelIndex) -> None:
        self.spin_frequency_center.setValue(self.model_found_lines.item(index.row(), 0))
        self.ensure_y_fits()

    def ensure_y_fits(self) -> None:
        if self._plot_line.xData is None or self._plot_line.xData.size < 2:
            return
        if self._plot_line.yData is None or self._plot_line.yData.size < 2:
            return
        x: pg.AxisItem = self._canvas.getAxis("bottom")
        y: pg.AxisItem = self._canvas.getAxis("left")
        visible_points: NDArray[np.float_] = self._plot_line.yData[
            (self._plot_line.xData >= min(x.range))
            & (self._plot_line.xData <= max(x.range))
        ]
        if np.any(visible_points < min(y.range)):
            minimum: np.float_ = np.min(visible_points)
            self.set_voltage_range(
                minimum - 0.05 * (max(y.range) - minimum), max(y.range)
            )
        if np.any(visible_points > max(y.range)):
            maximum: np.float_ = np.max(visible_points)
            self.set_voltage_range(
                min(y.range), maximum + 0.05 * (maximum - min(y.range))
            )

    def load_found_lines(self) -> None:
        def load_csv(fn: Path) -> Sequence[float]:
            sep: str = self.settings.csv_separator
            try:
                data: NDArray[np.float_] = (
                    np.loadtxt(
                        fn, delimiter=sep, usecols=(0,), encoding="utf-8", dtype=np.complex_
                    ).real
                    * 1e6
                )
            except ValueError:
                return []
            else:
                data = data[
                    (data >= self._plot_data.min_frequency)
                    & (data <= self._plot_data.max_frequency)
                ]
                return data

        def load_xlsx(fn: Path) -> Sequence[float]:
            from openpyxl.reader.excel import load_workbook
            from openpyxl.workbook.workbook import Workbook
            from openpyxl.worksheet.worksheet import Worksheet

            workbook: Workbook = load_workbook(
                fn, read_only=True, keep_vba=False, data_only=True
            )
            if len(workbook.sheetnames) != 1:
                return []
            sheet: Worksheet | None = workbook.active
            if sheet is None:
                return []

            data: list[float] = []
            reading_title: bool = True
            row: tuple[Any, ...]
            for row in sheet.values:
                if reading_title and isinstance(row[0], Number):
                    reading_title = False
                if not reading_title and not isinstance(row[0], Number):
                    break
                if not reading_title and isinstance(row[0], Number):
                    data.append(float(row[0]))
            if not data:
                return []

            data_: NDArray[np.float_] = np.asarray(data, dtype=np.float_) * 1e6
            data_ = data_[
                (data_ >= self._plot_data.min_frequency)
                & (data_ <= self._plot_data.max_frequency)
            ]
            return data_

        import importlib.util
        import mimetypes
        from itertools import chain

        mimetypes.init()

        supported_formats: dict[tuple[str, ...], str] = {
            tuple(all_cases(".csv")): _translate("file type", "Text with separators"),
            tuple(all_cases(".txt")): _translate("file type", "Plain text"),
        }
        supported_formats_callbacks: dict[str, Callable[[Path], Sequence[float]]] = {
            mimetypes.types_map[".csv"]: load_csv,
            mimetypes.types_map[".txt"]: load_csv,
        }
        if importlib.util.find_spec("openpyxl") is not None:
            supported_formats[tuple(all_cases(".xlsx"))] = _translate(
                "file type", "Microsoft Excel"
            )
            supported_formats_callbacks[mimetypes.types_map[".xlsx"]] = load_xlsx

        # reorder the dict
        supported_formats = {
            tuple(chain.from_iterable(supported_formats.keys())): _translate(
                "file type", "Supported formats"
            ),
            **supported_formats,
            (".*",): _translate("file type", "All files"),
        }

        filename, _filter = self.open_file_dialog(formats=supported_formats)
        if filename is None:
            return

        file_type: str | None = mimetypes.guess_type(filename)[0]
        if file_type is None or file_type not in supported_formats_callbacks:
            return
        new_lines: Sequence[float] = supported_formats_callbacks[file_type](filename)
        if not len(new_lines):
            return

        self.model_found_lines.add_lines(self._plot_data, new_lines)
        # add the new lines to the marked ones
        self.user_found_lines_data = np.concatenate(
            (self.user_found_lines_data, new_lines)
        )
        # avoid duplicates
        self.user_found_lines_data = self.user_found_lines_data[
            np.unique(self.user_found_lines_data, return_index=True)[1]
        ]
        # plot the data
        self.user_found_lines.setData(
            self.user_found_lines_data,
            self._plot_line.yData[
                self.model_found_lines.frequency_indices(
                    self._plot_data, self.user_found_lines_data
                )
            ],
        )
        self.toolbar.copy_trace_action.setEnabled(True)
        self.toolbar.save_trace_action.setEnabled(True)
        self.toolbar.clear_trace_action.setEnabled(True)

    def copy_found_lines(self) -> None:
        copy_to_clipboard(
            self.table_found_lines.stringify_table_plain_text(),
            self.table_found_lines.stringify_table_html(),
            Qt.TextFormat.RichText,
        )

    def save_found_lines(self) -> None:
        def save_csv(fn: Path) -> None:
            sep: str = self.settings.csv_separator
            with open(fn, "wt", encoding="utf-8") as f_out:
                f_out.writelines(
                    map(
                        lambda s: "# " + s + "\n",
                        [
                            sep.join(h.name for h in self.model_found_lines.header),
                            sep.join(h.unit for h in self.model_found_lines.header),
                        ],
                    )
                )
                for row in data:
                    f_out.write(
                        sep.join(
                            map(lambda x: str(x.real if x.imag == 0.0 else x), row)
                        )
                        + "\n"
                    )

        def save_xlsx(fn: Path) -> None:
            with pd.ExcelWriter(fn) as writer:
                df: pd.DataFrame = pd.DataFrame(data)
                df.to_excel(
                    writer,
                    index=False,
                    header=self.model_found_lines.header,
                    sheet_name=self._plot_line.name()
                    or _translate("workbook", "Sheet1"),
                )

        import importlib.util

        supported_formats: dict[tuple[str, ...], str] = {
            tuple(all_cases(".csv")): _translate("file type", "Text with separators")
        }
        supported_formats_callbacks: dict[str, Callable[[Path], None]] = {
            ".csv": save_csv
        }
        if importlib.util.find_spec("openpyxl") is not None:
            supported_formats[tuple(all_cases(".xlsx"))] = _translate(
                "file type", "Microsoft Excel"
            )
            supported_formats_callbacks[".xlsx"] = save_xlsx

        filename: Path | None
        _filter: str
        filename, _filter = self.save_file_dialog(formats=supported_formats)
        if filename is None:
            return

        f: NDArray[np.float_] = self.model_found_lines.all_data[:, 0] * 1e-6
        v: NDArray[np.float_] = self.model_found_lines.all_data[:, 1] * 1e3
        data: NDArray[np.complex_] | NDArray[np.float_]
        if self.model_found_lines.all_data.shape[1] > 2:
            g: NDArray[np.complex_] | NDArray[np.float_] = (
                self.model_found_lines.all_data[:, 2]
            )
            data = np.column_stack((f, v, g))
        else:
            data = np.column_stack((f, v))
        if np.all(data.imag == 0.0):
            data = data.real

        filename_ext: str = filename.suffix.casefold()
        if filename_ext in supported_formats_callbacks:
            supported_formats_callbacks[filename_ext](filename)

    def clear_automatically_found_lines(self) -> None:
        self.automatically_found_lines.clear()
        self.automatically_found_lines_data = np.empty(0)
        self._canvas.replot()

        self.model_found_lines.set_lines(self._plot_data, self.user_found_lines_data)
        self.toolbar.copy_trace_action.setEnabled(self.model_found_lines.is_empty)
        self.toolbar.save_trace_action.setEnabled(self.model_found_lines.is_empty)
        self.toolbar.clear_trace_action.setEnabled(self.model_found_lines.is_empty)
        self.button_clear_lines.setEnabled(False)
        self.button_next_line.setEnabled(False)
        self.button_prev_line.setEnabled(False)

    def clear_found_lines(self) -> None:
        self.automatically_found_lines.clear()
        self.automatically_found_lines_data = np.empty(0)
        self.user_found_lines.clear()
        self.user_found_lines_data = np.empty(0)
        self.model_found_lines.clear()
        self.toolbar.copy_trace_action.setEnabled(False)
        self.toolbar.save_trace_action.setEnabled(False)
        self.toolbar.clear_trace_action.setEnabled(False)
        self.button_clear_lines.setEnabled(False)
        self.button_next_line.setEnabled(False)
        self.button_prev_line.setEnabled(False)
        self._canvas.replot()

    def clear(self) -> None:
        close: QMessageBox = QMessageBox()
        close.setText(self.tr("Are you sure?"))
        close.setIcon(QMessageBox.Icon.Question)
        close.setWindowIcon(self.windowIcon())
        close.setWindowTitle(self.tr("Spectrometer Data Viewer"))
        close.setStandardButtons(
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
        )
        if close.exec() != QMessageBox.StandardButton.Yes:
            return

        self._ghost_line.clear()
        self._plot_line.clear()
        self._ghost_data.clear()
        self._plot_data.clear()
        self.clear_found_lines()
        self.toolbar.trace_action.setChecked(False)
        self.toolbar.clear_action.setEnabled(False)
        self.toolbar.open_ghost_action.setEnabled(False)
        self.toolbar.clear_ghost_action.setEnabled(False)
        self.toolbar.differentiate_action.setEnabled(False)
        self.toolbar.save_data_action.setEnabled(False)
        self.toolbar.copy_figure_action.setEnabled(False)
        self.toolbar.save_figure_action.setEnabled(False)
        self.toolbar.trace_action.setEnabled(False)
        self.toolbar.load_trace_action.setEnabled(False)
        self.toolbar.copy_trace_action.setEnabled(False)
        self.toolbar.save_trace_action.setEnabled(False)
        self.toolbar.clear_trace_action.setEnabled(False)
        self.box_find_lines.setEnabled(False)
        self._cursor_balloon.setVisible(False)
        self._crosshair_h_line.setVisible(False)
        self._crosshair_v_line.setVisible(False)
        self._cursor_x.setVisible(True)
        self._cursor_y.setVisible(True)
        self._canvas.replot()
        self.status_bar.clearMessage()
        self.setWindowTitle(self.tr("Spectrometer Data Viewer"))

    def clear_ghost(self) -> None:
        self._ghost_line.clear()
        self._ghost_data.clear()
        self.toolbar.clear_ghost_action.setEnabled(False)
        self._canvas.replot()

    def load_data(self, filename: Path | None = None) -> bool:
        self.clear_ghost()
        self.clear_found_lines()

        if filename is None:
            _filter: str
            _formats: dict[tuple[str, ...], str] = {
                (*all_cases(".conf"), *all_cases(".scandat")): _translate(
                    "file type", "PSK Spectrometer"
                ),
                tuple(all_cases(".fmd")): _translate(
                    "file type", "Fast Sweep Spectrometer"
                ),
            }
            filename, _filter = self.open_file_dialog(formats=_formats)
        if filename is None:
            return False

        v: NDArray[np.float_]
        f: NDArray[np.float_]
        g: NDArray[np.float_] = np.empty(0)
        jump: float
        if filename.suffix.casefold() == ".scandat":
            f, v, g, jump = load_data_scandat(filename, self)
            if f.size and v.size:
                self.settings.display_processing = True
                if jump > 0.0:
                    self._data_mode = self.PSK_WITH_JUMP_DATA_MODE
                else:
                    self._data_mode = self.PSK_DATA_MODE
        elif filename.suffix.casefold() in (".csv", ".conf"):
            f, v, g, jump = load_data_csv(filename)
            if f.size and v.size:
                self.settings.display_processing = True
                if jump > 0.0:
                    self._data_mode = self.PSK_WITH_JUMP_DATA_MODE
                else:
                    self._data_mode = self.PSK_DATA_MODE
        elif filename.suffix.casefold() in (".fmd", ".frd"):
            f, v = load_data_fs(filename)
            if f.size and v.size:
                self.settings.display_processing = False
                self._data_mode = self.FS_DATA_MODE
        else:
            return False

        if not (f.size and v.size):
            return False

        if self._data_mode == self.FS_DATA_MODE:
            self.switch_data_action.setChecked(False)

        self._plot_line.setData(
            f,
            (g if self.switch_data_action.isChecked() else v),
            name=str(filename.parent / filename.stem),
        )
        self._plot_data.set_data(frequency_data=f, gamma_data=g, voltage_data=v)

        min_frequency: np.float_ = f[0]
        max_frequency: np.float_ = f[-1]

        self.toolbar.clear_action.setEnabled(True)
        step: int = int(
            round(self.settings.jump / ((max_frequency - min_frequency) / (f.size - 1)))
        )
        self.toolbar.differentiate_action.setEnabled(
            self._data_mode == self.PSK_DATA_MODE and 0 < step < 0.25 * f.size
        )
        self.switch_data_action.setEnabled(
            self._data_mode in (self.PSK_DATA_MODE, self.PSK_WITH_JUMP_DATA_MODE)
        )
        self.toolbar.save_data_action.setEnabled(True)
        self.toolbar.copy_figure_action.setEnabled(True)
        self.toolbar.save_figure_action.setEnabled(True)
        self.toolbar.trace_action.setEnabled(True)
        self.box_find_lines.setEnabled(bool(self.model_signal.size))

        self._loading = True
        self.spin_frequency_min.setMaximum(
            max(max_frequency, self.spin_frequency_min.value())
        )
        self.spin_frequency_max.setMinimum(
            min(min_frequency, self.spin_frequency_max.value())
        )
        if not self.check_frequency_persists.isChecked():
            self.spin_frequency_min.setValue(min_frequency)
            self.spin_frequency_max.setValue(max_frequency)
            self.spin_frequency_span.setValue(max_frequency - min_frequency)
            self.spin_frequency_center.setValue(0.5 * (max_frequency + min_frequency))
        self._loading = False

        self.display_gamma_or_voltage()

        self.set_frequency_range(
            lower_value=self.spin_frequency_min.value(),
            upper_value=self.spin_frequency_max.value(),
        )
        self.set_voltage_range(
            lower_value=self.spin_voltage_min.value(),
            upper_value=self.spin_voltage_max.value(),
        )

        self.setWindowTitle(self.tr("%s — Spectrometer Data Viewer") % filename)

        self.toolbar.open_ghost_action.setEnabled(True)

        return True

    def load_ghost_data(self, filename: Path | None = None) -> bool:
        if filename is None:
            _filter: str
            _formats: dict[tuple[str, ...], str] = {
                (*all_cases(".conf"), *all_cases(".scandat")): _translate(
                    "file type", "PSK Spectrometer"
                ),
                tuple(all_cases(".fmd")): _translate(
                    "file type", "Fast Sweep Spectrometer"
                ),
            }
            filename, _filter = self.open_file_dialog(formats=_formats)
        if filename is None:
            return False

        v: NDArray[np.float_]
        f: NDArray[np.float_]
        g: NDArray[np.float_] = np.empty(0)
        jump: float
        if filename.suffix.casefold() == ".scandat":
            f, v, g, jump = load_data_scandat(filename, self)
            if f.size and v.size:
                if jump > 0.0:
                    if self._data_mode != self.PSK_WITH_JUMP_DATA_MODE:
                        return False
                else:
                    if self._data_mode != self.PSK_DATA_MODE:
                        return False
        elif filename.suffix.casefold() in (".csv", ".conf"):
            f, v, g, jump = load_data_csv(filename)
            if f.size and v.size:
                if jump > 0.0:
                    if self._data_mode != self.PSK_WITH_JUMP_DATA_MODE:
                        return False
                else:
                    if self._data_mode != self.PSK_DATA_MODE:
                        return False
        elif filename.suffix.casefold() in (".fmd", ".frd"):
            f, v = load_data_fs(filename)
            if f.size and v.size:
                if self._data_mode != self.FS_DATA_MODE:
                    return False
        else:
            return False

        if not (f.size and v.size):
            return False

        self._ghost_line.setData(
            f,
            (g if self.switch_data_action.isChecked() else v),
            name=str(filename.parent / filename.stem),
        )
        self._ghost_data.set_data(frequency_data=f, gamma_data=g, voltage_data=v)

        self.toolbar.clear_ghost_action.setEnabled(True)

        self.display_gamma_or_voltage()

        return True

    @property
    def trace_mode(self) -> bool:
        return self.toolbar.trace_action.isChecked()

    def actions_off(self) -> None:
        self.toolbar.trace_action.setChecked(False)

    def calculate_second_derivative(self) -> None:
        self._data_mode = self.PSK_WITH_JUMP_DATA_MODE
        self.display_gamma_or_voltage()
        self.model_found_lines.refresh()

    def on_switch_data_action_toggled(self, new_state: bool) -> None:
        self._plot_data.data_type = (
            PlotDataItem.GAMMA_DATA if new_state else PlotDataItem.VOLTAGE_DATA
        )
        self._ghost_data.data_type = (
            PlotDataItem.GAMMA_DATA if new_state else PlotDataItem.VOLTAGE_DATA
        )
        self.set_config_value("display", "unit", self._plot_data.data_type)
        self.display_gamma_or_voltage(new_state)

    def display_gamma_or_voltage(self, display_gamma: bool | None = None) -> None:
        if display_gamma is None:
            display_gamma = self.switch_data_action.isChecked()

        if self.toolbar.differentiate_action.isChecked():
            self._plot_data.jump = self.settings.jump
            self._ghost_data.jump = self.settings.jump
        else:
            self._plot_data.jump = np.nan
            self._ghost_data.jump = np.nan

        if display_gamma:
            self.box_voltage.setWindowTitle(self.tr("Absorption"))
        else:
            self.box_voltage.setWindowTitle(self.tr("Voltage"))

        if self._plot_data:  # something is loaded
            self._plot_line.setData(self._plot_data.x_data, self._plot_data.y_data)

            self._loading = True
            y_data: NDArray[np.float_] = self._plot_data.y_data
            min_y: np.float_ = np.min(y_data)
            max_y: np.float_ = np.max(y_data)
            if not self.check_voltage_persists.isChecked():
                self.on_ylim_changed((min_y, max_y))
            self.spin_voltage_min.setMaximum(max(max_y, self.spin_voltage_min.value()))
            self.spin_voltage_max.setMinimum(min(min_y, self.spin_voltage_max.value()))
            self._loading = False

        if self._ghost_data:  # something is loaded
            self._ghost_line.setData(self._ghost_data.x_data, self._ghost_data.y_data)

        if self.automatically_found_lines_data.size:  # something is marked
            self.automatically_found_lines.setData(
                self.automatically_found_lines_data,
                self._plot_data.y_data[
                    self.model_found_lines.frequency_indices(
                        self._plot_data, self.automatically_found_lines_data
                    )
                ],
            )
        if self.user_found_lines_data.size:  # something is marked
            self.user_found_lines.setData(
                self.user_found_lines_data,
                self._plot_data.y_data[
                    self.model_found_lines.frequency_indices(
                        self._plot_data, self.user_found_lines_data
                    )
                ],
            )

        a: pg.AxisItem = self._canvas.getAxis("left")
        if display_gamma:
            self.check_voltage_persists.setText(self.tr("Keep absorption range"))

            a.enableAutoSIPrefix(False)
            a.setLabel(
                text=_translate("plot axes labels", "Absorption"),
                units=_translate("unit", "cm<sup>−1</sup>"),
            )
            a.scale = 1.0
            a.autoSIPrefixScale = 1.0

            self._cursor_y.suffix = _translate("unit", "cm<sup>−1</sup>")
            self._cursor_y.siPrefix = False
            self._cursor_y.setFormatStr(
                "{mantissa:.{decimals}f}×10<sup>{exp}</sup>{suffixGap}{suffix}"
            )
            opts = {
                "suffix": _translate("unit", "cm⁻¹"),
                "siPrefix": False,
                "format": "{value:.{decimals}e}{suffixGap}{suffix}",
            }

        else:
            self.check_voltage_persists.setText(self.tr("Keep voltage range"))

            a.enableAutoSIPrefix(True)
            a.setLabel(
                text=_translate("plot axes labels", "Voltage"),
                units=_translate("unit", "V"),
            )

            self._cursor_y.suffix = _translate("unit", "V")
            self._cursor_y.siPrefix = True
            self._cursor_y.setFormatStr(
                "{scaledValue:.{decimals}f}{suffixGap}{siPrefix}{suffix}"
            )
            opts = {
                "suffix": _translate("unit", "V"),
                "siPrefix": True,
                "format": "{scaledValue:.{decimals}f}{suffixGap}{siPrefix}{suffix}",
            }
        self.spin_voltage_min.setOpts(**opts)
        self.spin_voltage_max.setOpts(**opts)

        self.hide_cursors()

    def save_data(self) -> None:
        if self._plot_line.yData is None:
            return

        def save_csv(fn: Path) -> None:
            data: NDArray[np.float_]
            sep: str = self.settings.csv_separator
            if self.switch_data_action.isChecked():
                data = np.column_stack((x * 1e-6, y))
                # noinspection PyTypeChecker
                np.savetxt(
                    fn,
                    data,
                    delimiter=sep,
                    header=(
                        sep.join(
                            (
                                _translate("plot axes labels", "Frequency"),
                                _translate("plot axes labels", "Absorption"),
                            )
                        )
                        + "\n"
                        + sep.join(
                            (_translate("unit", "MHz"), _translate("unit", "cm⁻¹"))
                        )
                    ),
                    fmt=("%.3f", "%.6e"),
                    encoding="utf-8",
                )
            else:
                data = np.column_stack((x * 1e-6, y * 1e3))
                # noinspection PyTypeChecker
                np.savetxt(
                    filename,
                    data,
                    delimiter=sep,
                    header=(
                        sep.join(
                            (
                                _translate("plot axes labels", "Frequency"),
                                _translate("plot axes labels", "Voltage"),
                            )
                        )
                        + "\n"
                        + sep.join(
                            (_translate("unit", "MHz"), _translate("unit", "mV"))
                        )
                    ),
                    fmt=("%.3f", "%.6f"),
                    encoding="utf-8",
                )

        def save_xlsx(fn: Path) -> None:
            data: NDArray[np.float_]
            with pd.ExcelWriter(fn) as writer:
                df: pd.DataFrame
                if self.switch_data_action.isChecked():
                    data = np.column_stack((x * 1e-6, y))
                    df = pd.DataFrame(data)
                    df.to_excel(
                        writer,
                        index=False,
                        header=[
                            self.model_found_lines.header[0],
                            self.model_found_lines.header[2],
                        ],
                        sheet_name=self._plot_line.name()
                        or _translate("workbook", "Sheet1"),
                    )
                else:
                    data = np.column_stack((x * 1e-6, y * 1e3))
                    df = pd.DataFrame(data)
                    df.to_excel(
                        writer,
                        index=False,
                        header=[
                            self.model_found_lines.header[0],
                            self.model_found_lines.header[1],
                        ],
                        sheet_name=self._plot_line.name()
                        or _translate("workbook", "Sheet1"),
                    )

        import importlib.util

        supported_formats: dict[tuple[str, ...], str] = {
            tuple(all_cases(".csv")): _translate("file type", "Text with separators")
        }
        supported_formats_callbacks: dict[str, Callable[[Path], None]] = {
            ".csv": save_csv
        }
        if importlib.util.find_spec("openpyxl") is not None:
            supported_formats[tuple(all_cases(".xlsx"))] = _translate(
                "file type", "Microsoft Excel"
            )
            supported_formats_callbacks[".xlsx"] = save_xlsx

        filename: Path | None
        _filter: str
        filename, _filter = self.save_file_dialog(formats=supported_formats)
        if filename is None:
            return
        x: NDArray[np.float_] = self._plot_line.xData
        y: NDArray[np.float_] = self._plot_line.yData
        max_mark: float
        min_mark: float
        min_mark, max_mark = self._canvas.axes["bottom"]["item"].range
        good: NDArray[np.bool_] = (min_mark <= x) & (x <= max_mark)
        x = x[good]
        y = y[good]
        del good

        filename_ext: str = filename.suffix.casefold()
        if filename_ext in supported_formats_callbacks:
            supported_formats_callbacks[filename_ext](filename)

    def copy_figure(self) -> None:
        exporter: ImageExporter = ImageExporter(self._canvas)
        self.hide_cursors()
        exporter.export(copy=True)

    def save_figure(self) -> None:
        exporter: ImageExporter = ImageExporter(self._canvas)
        formats: dict[tuple[str, ...], str] = {
            tuple(exporter.getSupportedImageFormats()): _translate(
                "file dialog", "Image files"
            )
        }
        filename: Path | None
        _filter: str
        filename, _filter = self.save_file_dialog(formats=formats)
        if filename is None:
            return
        self.hide_cursors()
        exporter.export(filename)
